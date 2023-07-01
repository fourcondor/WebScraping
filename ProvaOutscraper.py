# AIzaSyCJEQcHyw6u1llI678mO5BlsKCNAYr7kGQ
from googleapiclient.discovery import build
import json
from openpyxl import Workbook
import csv
import pprint

# Inserisci la tua chiave API qui
API_KEY = 'AIzaSyCJEQcHyw6u1llI678mO5BlsKCNAYr7kGQ'

# Crea l'oggetto di servizio per l'API di ricerca di Google
service = build('customsearch', 'v1', developerKey=API_KEY)

# Definisci il tuo ID di ricerca personalizzato
# Puoi crearne uno nell'interfaccia di amministrazione delle Ricerche personalizzate di Google
CUSTOM_SEARCH_ENGINE_ID = '85f179deb03424abc'

# Query di ricerca
lista = []
count = 0
categoria = 'guarnizioni moto'
estensioni = ['shop on line', 'vendita on line', 'acquista on line', 'ecommerce', 'e-commerce']
# for estensione in estensioni:
#     query = f'{categoria} {estensione} -site:amazon.* -site:ebay.* -site:trovaprezzi.it -site:collistar.it -site:idealo.it -inurl:\"category/sponsored-post\"'
#
#     # Esegui la ricerca utilizzando l'API di Google
#     result = service.cse().list(q=query, cx=CUSTOM_SEARCH_ENGINE_ID, num=10).execute()
#
#     # Estrai i risultati dalla risposta
#     items = result['items']
#     print(result)
#
#     # Stampa i titoli e i link dei risultati
#     for item in items:
#         count += 1
#         print('Title: {}'.format(item['title']))
#         print('Link: {}'.format(item['link']))
#         lista.append(item['link'])
#         print('---')
#
# print(f"COUNT LINKS {count}")
#
# # Percorso del file CSV
# file_path = './risultati.csv'
#
# # Apri il file CSV in modalità scrittura
# with open(file_path, mode='w', newline='') as file:
#     # Crea un writer CSV
#     writer = csv.writer(file)
#
#     # Scrivi la lista di risultati nel file CSV
#     writer.writerows([[link] for link in lista])
#
# print(f"I risultati sono stati scritti nel file '{file_path}'")
#
#
#
# # Vettore di link da scrivere nel file Excel
# links = ['link1', 'link2', 'link3']
#
# # Percorso del file Excel
# file_path = f'./{categoria}.xlsx'
#
# # Crea un nuovo workbook Excel
# workbook = Workbook()
#
# # Seleziona il foglio di lavoro predefinito
# sheet = workbook.active
#
# # Scrivi i link su righe separate nel file Excel
# for index, link in enumerate(lista, start=1):
#     sheet.cell(row=index, column=2, value=link)
#
# # Salva il workbook nel file Excel
# workbook.save(file_path)
#
# print(f"I link sono stati scritti nel file '{file_path}'")
# Prova JSON in excel
query = f'{categoria} shop on line -site:amazon.* -site:ebay.* -site:trovaprezzi.it -site:collistar.it -site:idealo.it -inurl:\"category/sponsored-post\"'
# Esegui la ricerca utilizzando l'API di Google
result = service.cse().list(q=query, cx=CUSTOM_SEARCH_ENGINE_ID, num=10).execute()
# Percorso del file Excel
file_path = './JSON_GOOGLE.xlsx'



# Crea un nuovo workbook Excel
workbook = Workbook()

# Seleziona il foglio di lavoro predefinito
sheet = workbook.active

# Funzione per scrivere un dizionario nel foglio di lavoro
def write_dict_to_sheet(dictionary, start_row):
    if len(dictionary) == 0:
        return

    # Scrivi le chiavi come intestazioni di colonna
    keys = list(dictionary.keys())
    num_columns = len(keys)
    for col in range(num_columns):
        sheet.cell(row=start_row, column=col + 1, value=keys[col])

    # Serializza i valori JSON in formato stringa
    values = [json.dumps(value) for value in dictionary.values()]

    # Scrivi i valori nel foglio di lavoro
    sheet.append(values)

# Scrivi il risultato nel foglio di lavoro
write_dict_to_sheet(result, 1)

# Salva il workbook nel file Excel
workbook.save(file_path)

print(f"I risultati sono stati scritti nel file '{file_path}'")
