import csv
import openpyxl
import os
import scraping
import scrapingAmazon


def aggiorna_report_ecommerce(nome_prodotto, lista_ecommerce):
    # Percorso del file Excel
    excel_file_path = 'ReportProspetto.xlsx'
    intestazione = ['Nome Prodotto', 'PROSPECT']

    # Verifica se il file esiste
    if os.path.isfile(excel_file_path):
        # Il file esiste, apri il file Excel in modalità di append
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        print("FILE ESISTENTE")
    else:
        # Il file non esiste, crea un nuovo file Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        print("CREO FILE")
        last_row = worksheet.max_row
        for col, value in enumerate(intestazione, start=1):
            worksheet.cell(row=last_row, column=col).value = value

    # Riga in cui scrivere i nuovi dati
    #nome prodotto lista[0]
    #               lista[1]
    #               lista[2]
    #               ecc..
    step = 0
    for i in range(0, len(lista_ecommerce)):
        if i == 0:
            new_row = [nome_prodotto, lista_ecommerce[i]]
        else:
            new_row = ['', lista_ecommerce[i]]
            if lista_ecommerce[i] == '':
                continue
        # Trova l'ultima riga utilizzata nel foglio di lavoro
        last_row = worksheet.max_row + step + 1

        # Scrivi i nuovi dati nella riga successiva
        for col, value in enumerate(new_row, start=1):
            worksheet.cell(row=last_row, column=col).value = value
        step += 1

    # Salva le modifiche nel file Excel
    workbook.save(excel_file_path)
    # Formatta file Excel
    scraping.formatta_file_excel('ReportProspetto.xlsx')
    print("//////REPORT PROSPETTO AGGIORNATO!/////")
