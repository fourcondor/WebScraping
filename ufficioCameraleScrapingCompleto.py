import requests
import openpyxl
import os
import sys

# COSTANTI
api_url = 'https://imprese.openapi.it/advance/{}'

# Partita IVA, codice fiscale o ID azienda di interesse
# piva_cf_or_id = '02080418021'

# Token di autenticazione
token = '649b05c3165cce787f26377e'

# 649acbb9165cce787f263752

headers = {
    'Authorization': f'Bearer {token}'
}

file_path = ""

# TRACCIATO
tracciato = {'data_denominazione': None,

             'data_cf': None,
             'data_piva': None,
             'data_dettaglio_bilanci_2021_fatturato': None,
             'data_dettaglio_bilanci_2022_fatturato': None,
             'data_dettaglio_bilanci_2020_dipendenti': None,
             'data_dettaglio_bilanci_2021_dipendenti': None,
             'data_dettaglio_bilanci_2022_dipendenti': None,
             'data_dettaglio_bilanci_2023_dipendenti': None,
             'data_stato_attivita': None,
             'data_dettaglio_cciaa': None,
             'data_dettaglio_codice_ateco': None,
             'data_dettaglio_descrizione_ateco': None,
             'data_dettaglio_data_inizio_attivita': None,
             'data_dettaglio_pec': None,
             'data_dettaglio_soci_0_denominazione': None,
             'data_dettaglio_soci_0_nome': None,
             'data_dettaglio_soci_0_cognome': None,
             'data_dettaglio_soci_0_cf_socio': None,
             'data_gps_coordinates_0': None,
             'data_gps_coordinates_1': None,
             'Categoria': None,
             }


def flatten_json(json_data, parent_key='', flattened_dict=None):
    if flattened_dict is None:
        flattened_dict = {}
    if isinstance(json_data, dict):
        for key, value in json_data.items():
            new_key = parent_key + '_' + key if parent_key else key
            if isinstance(value, dict):
                flatten_json(value, new_key, flattened_dict)
            elif isinstance(value, list):
                for i, item in enumerate(value):
                    new_item_key = new_key + '_' + str(i)
                    if isinstance(item, dict):
                        flatten_json(item, new_item_key, flattened_dict)
                    else:
                        flattened_dict[new_item_key] = item
            else:
                flattened_dict[new_key] = value
    return flattened_dict


def aggiorna_excel():
    file_path_excel = "piva.xlsx"
    if os.path.isfile(file_path_excel):
        # Il file esiste, apri il file Excel
        workbook = openpyxl.load_workbook(file_path_excel)
        worksheet = workbook.active
    else:
        # Il file non esiste, crea un nuovo file Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Inserisci la prima riga con i nomi delle chiavi
        for col, key in enumerate(tracciato.keys(), start=1):
            worksheet.cell(row=1, column=col).value = key

    # Trova la prossima riga vuota nel foglio di lavoro
    next_row = worksheet.max_row + 1
    # Inserisci i valori delle chiavi
    for col, key in enumerate(tracciato.keys(), start=1):
        worksheet.cell(row=next_row, column=col).value = tracciato[key]

    # Salva il file Excel
    workbook.save(file_path_excel)
    return


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Specificare il nome del file come parametro.")
        sys.exit(1)

    file_path = sys.argv[1]

    print("Il nome del file inserito è:", file_path)
    print("Indicare categoria : \n(1) amazon\n(2) ebay")
    choice = input()
    while int(choice) != 1 and int(choice) != 2:
        print("Categoria scelta errata o non presente nell'elenco")
        choice = input()
    if int(choice) == 1:
        categoria = 'Amazon'
    if int(choice) == 2:
        categoria = 'Ebay'
    print("Estrazione Dati")
    piva_cf_or_id_list = []
    with open(file_path, 'r') as file:
        content = file.read()
        multi_IVA = content.replace("\n", " ").split()
        for iva in multi_IVA:
            piva_cf_or_id_list.append(iva)
    for piva_cf_or_id in piva_cf_or_id_list:
        url = api_url.format(piva_cf_or_id)
        response = requests.get(url, headers=headers)
        # Ottieni la risposta dell'API
        api_response = response.json()
        #  ESTRAGGO I DATI DAL JSON E IMPORTO SU EXCEL
        dati_excel = flatten_json(api_response)
        print(dati_excel)
        for cell in tracciato.keys():
            if cell in dati_excel.keys():
                tracciato[cell] = dati_excel[cell]
            else:
                tracciato[cell] = 'Non disponibile'
        tracciato['Categoria'] = categoria
        print(f"{'-' * 20}")
        print(tracciato)
        # CONVERTO JSON IN EXCEL
        aggiorna_excel()
